# -*- coding: utf-8 -*-

import re

from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.cell_range import CellRange

def getTableObject(ws, tableName):

    # Определение нужной таблицы
    table = None
    for tbl in ws._tables:
        if tbl.displayName == str(tableName):
            table = tbl
            break
    return table

def addTableRows(wb, tableName, rowsCount):
    ws = wb.active

    # Определение диапазона таблицы
    table = getTableObject(ws, tableName)
    initRange = CellRange(table.ref)

    # Определение диапазона, который включает в себя всё, что находится ниже данной таблицы
    rangeToMove = CellRange(
        min_col=initRange.min_col,
        max_col=initRange.max_col,
        min_row=initRange.max_row + 1,
        max_row=ws.max_row
    )

    # Определение именованных диапазонов, находящихся на активном листе ниже данной таблицы
    definedNames = {}
    for name in wb.defined_names.definedName:
        for dest in name.destinations:
            if (dest[0] == ws.title) and (not CellRange(dest[1]).isdisjoint(rangeToMove)):
                if name.name in definedNames.keys():
                    definedNames[name.name].append(dest[1])
                else:
                    definedNames[name.name] = [dest[1]]

    # Переопределение всех объединённых ячеек, находящихся ниже данной таблицы
    mergedCells = [a for a in ws.merged_cells.ranges]
    for item in mergedCells:
        if not item.isdisjoint(rangeToMove):
            ws.unmerge_cells(str(item))
            item.shift(row_shift=rowsCount)
            ws.merge_cells(str(item))

    # Дополнительное перемещение именованных диапазонов, находящихся ниже данной таблицы
    for name, rangeList in definedNames.items():
        newRangeList = []
        for range in rangeList:
            rangeObject = CellRange(range)
            rangeObject.shift(row_shift=rowsCount)
            newRangeList.append(re.sub(r'([A-Z]+|[0-9]+)', r'$\1', str(rangeObject)))
        newNamedRange = DefinedName(
            name=name,
            attr_text='%s!%s' %  (ws.title, ';'.join(newRangeList))
        )
        del wb.defined_names[name]
        wb.defined_names.append(newNamedRange)

    # Дополнительное перемещение таблиц, находящихся ниже данной таблицы
    for tbl in ws._tables:
        if not CellRange(tbl.ref).isdisjoint(rangeToMove):
            newTblRange = CellRange(tbl.ref)
            newTblRange.shift(row_shift=rowsCount)
            tbl.ref = str(newTblRange)

    # Сдвиг rangeToMove на количество добавляемых в таблицу строк
    ws.move_range(rangeToMove, rows=rowsCount)
    newTableRange = initRange
    newTableRange.expand(down=rowsCount)
    table.ref = str(newTableRange)