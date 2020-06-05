import sys
import os
import openpyxl
import MySQLdb
from PyQt5 import QtWidgets

import defsexcel
from forms import mainwindow, checkdataform


class MyWindow(QtWidgets.QMainWindow, mainwindow.Ui_MainWindow):
    projectData = {}
    finalLayerDict = {}
    finalPairDict = {}

    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.dbMpwRS = None
        self.dbTzRS = None
        self.dbBaseRouteRS = None
        self.dbOptionRS = None
        self.dbBaseLayerRS = None
        self.dbOptionLayerRS = None
        self.checkDataForm = None

        self.listWidget.setSelectionMode(2)

        conn = MySQLdb.connect('localhost', 'root', '', 'test01', charset='utf8')
        self.dbSelect = conn.cursor()

        self.dbSelect.execute("""
            SELECT DISTINCT Technology.name, Technology.id, Technology.shortname
            FROM Technology LEFT JOIN BaseRoute
            ON Technology.Id = BaseRoute.TechnologyId
            WHERE BaseRoute.RouteStatus = 2
        """)
        self.dbTechRS = self.dbSelect.fetchall()
        for line in self.dbTechRS:
            self.comboBox.addItem(line[0])
        self.comboBox.setCurrentIndex(-1)

        self.comboBox_4.addItem("SPW")
        self.comboBox_4.addItem("MPW")
        self.comboBox_4.setCurrentIndex(-1)
        self.comboBox_5.addItem("SLR")
        self.comboBox_5.addItem("MLR")
        self.comboBox_5.setCurrentIndex(-1)
        self.checkBox.setEnabled(False)

        self.comboBox.activated.connect(self.techFamilyCmbChange)
        self.comboBox_2.activated.connect(self.projNameCmbChange)
        self.comboBox_5.activated.connect(self.maskTypeCmbChange)
        self.radioButton.toggled.connect(self.withoutTegBtClick)
        self.pushButton.clicked.connect(sys.exit)
        self.pushButton_2.clicked.connect(self.okBtnClick)

    def techFamilyCmbChange(self):
        chosenTechName = self.dbTechRS[self.comboBox.currentIndex()][0]
        chosenTechId = self.dbTechRS[self.comboBox.currentIndex()][1]
        MyWindow.techCode = self.dbTechRS[self.comboBox.currentIndex()][2]

        self.comboBox_2.clear()
        self.comboBox_3.clear()
        self.listWidget.clear()

        self.dbSelect.execute("""
          SELECT DISTINCT `UI::MPW`.Id, `UI::MPW`.Name, `UI::MPW`.MaskSetNumber, `UI::MPW`.RegisterId
          FROM `UI::MPW` LEFT JOIN `UI::TZ`
          ON `UI::MPW`.Id = `UI::TZ`.MpwId
          WHERE `UI::TZ`.StatusDevelopmentId = 1
          AND `UI::MPW`.TechnologyId = """ + str(chosenTechId))
        self.dbMpwRS = self.dbSelect.fetchall()
        if not self.dbMpwRS:
            QtWidgets.QMessageBox.information(self, "Информация", "Не найдено ни одного доступного запуска.\nВыберите "
                                                                  "другое семейство технологий",
                                              QtWidgets.QMessageBox.Ok)
        else:
            for line in self.dbMpwRS:
                self.comboBox_2.addItem(line[1])
        self.comboBox_2.setCurrentIndex(-1)

        self.dbSelect.execute("""
        SELECT `Technology::Teg`.Path 
        FROM `Technology::Teg`
        WHERE `Technology::Teg`.TechnologyId = """ + str(chosenTechId) + """
        AND `Technology::Teg`.Lock = 0
        """)
        dbTegListRS = self.dbSelect.fetchall()
        if not dbTegListRS:
            self.listWidget.addItem('Тестовые сборки не найдены')
        else:
            for line in dbTegListRS:
                self.listWidget.addItem(str(line[0])[str(line[0]).rfind('/') + 1:])

    def projNameCmbChange(self):
        self.comboBox_3.clear()
        lineText = self.comboBox_2.currentText()  # проверить регулярным выражением!
        chosenProjName = self.comboBox_2.currentText()
        chosenProjId = self.dbMpwRS[self.comboBox_2.currentIndex()][0]

        self.dbSelect.execute("""
                  SELECT `UI::TZ`.Id, `UI::TZ`.Name
                  FROM `UI::TZ`
                  WHERE `UI::TZ`.MpwId = """ + str(chosenProjId) + """
                  AND `UI::TZ`.StatusDevelopmentId = 1
        """)
        self.dbTzRS = self.dbSelect.fetchall()

        if not self.dbTzRS:
            QtWidgets.QMessageBox.critical(self, "Ошибочка", "Не найдено ни одного доступного ТЗ для существующего "
                                                             "запуска.\nТакого быть не должно, обратитесь к "
                                                             "смотрящему за базой данных.",
                                           QtWidgets.QMessageBox.Ok)
        else:
            for line in self.dbTzRS:
                self.comboBox_3.addItem(line[1])  # без сортировки
        self.comboBox_3.setCurrentIndex(-1)

    def maskTypeCmbChange(self):
        self.checkBox.setCheckState(False)
        if self.comboBox_5.currentText() == 'SLR':
            self.checkBox.setEnabled(False)
        else:
            self.checkBox.setEnabled(True)

    def withoutTegBtClick(self):
        if self.radioButton_2.isChecked():
            self.listWidget.clearSelection()
            self.listWidget.setEnabled(False)
        else:
            self.listWidget.setEnabled(True)

    def initFormValidator(self):
        errList = []

        # ЁЩЁ ПРОВЕРИТЬ РЕГУЛЯРНЫМ ВЫРАЖЕНИЕМ!

        if self.comboBox.currentIndex() == -1:
            errList.append('-- Семейство технологий (не указано)')

        if self.comboBox_2.currentIndex() == -1:
            errList.append('-- Название запуска (не найден в базе данных)')

        if self.comboBox_3.currentIndex() == -1:
            errList.append('-- Версия ТЗ (не найдено в базе)')

        if self.comboBox_4.currentIndex() == -1:
            errList.append('-- Тип запуска (не указан)')

        if self.comboBox_5.currentIndex() == -1:
            errList.append('-- Тип шаблона (не указан)')

        if not self.listWidget.selectedItems() and self.radioButton.isChecked():
            errList.append('-- Тестовые сборки: флаг включен, а сборки не выбраны')

        if errList:
            errString = 'Неверно указаны входные параметры:\n'
            for item in errList:
                errString = errString + item + '\n'
            return errString

    def okBtnClick(self):
        if self.initFormValidator():
            QtWidgets.QMessageBox.critical(self, "Ошибочка", self.initFormValidator(), QtWidgets.QMessageBox.Ok)
        else:
            self.projectData['techId'] = str(self.dbTechRS[self.comboBox.currentIndex()][1])
            self.projectData['techFamily'] = str(self.comboBox.currentText())
            self.projectData['techShortName'] = str(self.dbTechRS[self.comboBox.currentIndex()][2])
            self.projectData['projectName'] = str(self.comboBox_2.currentText())
            self.projectData['tzVer'] = str(self.comboBox_3.currentText())
            self.projectData['projectType'] = str(self.comboBox_4.currentText())
            self.projectData['maskType'] = str(self.comboBox_5.currentText())
            self.projectData['maskSetNum'] = str(self.dbMpwRS[self.comboBox_2.currentIndex()][2])

            regId = int(self.dbMpwRS[self.comboBox_2.currentIndex()][3])
            self.dbSelect.execute("""
                      SELECT * FROM `UI::MPWRegisterList`
                      WHERE `UI::MPWRegisterList`.Id = 
                    """ + str(regId))
            dbRegRS = self.dbSelect.fetchone()
            if not dbRegRS:
                self.projectData['registerChar'] = 'Не обнаружен'
            else:
                self.projectData['registerChar'] = str(dbRegRS[2])
                self.projectData['registerName'] = str(dbRegRS[1])
            self.projectData['tzId'] = str(self.dbTzRS[self.comboBox_3.currentIndex()][0])

            self.dbSelect.execute("""
                              SELECT DISTINCT `BaseRoute`.Id, `BaseRoute`.Name FROM `UI::TZ`
                              LEFT JOIN  `UI::TZ_UI::Application_Relation`
                              ON `UI::TZ`.Id = `UI::TZ_UI::Application_Relation`.TzId
                              LEFT JOIN `UI::Application`
                              ON `UI::TZ_UI::Application_Relation`.ApplicationId = `UI::Application`.Id
                              LEFT JOIN `Storage::Incheck`
                              ON `UI::Application`.StorageIncheckId = `Storage::Incheck`.Id
                              LEFT JOIN `BaseRoute`
                              ON `Storage::Incheck`.BaseRouteId = `BaseRoute`.Id
                              WHERE `UI::TZ`.Id = """ + str(self.projectData['tzId']) + """
                              AND `UI::Application`.Lock = 0
                            """)
            self.dbBaseRouteRS = self.dbSelect.fetchall()

            baseRouteArr = []
            if self.dbBaseRouteRS:
                for line in self.dbBaseRouteRS:
                    baseRouteArr.append(str(line[1]))
            else:
                baseRouteArr.append('Не определены')
            self.projectData['baseRoutes'] = baseRouteArr

            self.dbSelect.execute("""
                              SELECT DISTINCT `MaskLayer`.Id, `Option`.Name FROM `UI::TZ`
                              LEFT JOIN  `UI::TZ_UI::Application_Relation`
                              ON `UI::TZ`.Id = `UI::TZ_UI::Application_Relation`.TzId
                              LEFT JOIN `UI::Application`
                              ON `UI::TZ_UI::Application_Relation`.ApplicationId = `UI::Application`.Id
                              LEFT JOIN `Storage::Incheck`
                              ON `UI::Application`.StorageIncheckId = `Storage::Incheck`.Id
                              LEFT JOIN `OptionCombination_Option_Relation`
                              ON `Storage::Incheck`.OptionCombinationId = `OptionCombination_Option_Relation`.OptionCombinationId
                              LEFT JOIN `Option`
                              ON `OptionCombination_Option_Relation`.OptionId = `Option`.Id
                              LEFT JOIN `MaskLayer_Option_Relation`
                              ON `Option`.Id = `MaskLayer_Option_Relation`.OptionId
                              LEFT JOIN `MaskLayer`
                              ON `MaskLayer_Option_Relation`.MaskLayerId = `MaskLayer`.Id
                              WHERE `UI::Application`.Lock = 0
                              AND `Storage::Incheck`.OptionCombinationId IS NOT NULL
                              AND `Option`.BaseRouteId = `Storage::Incheck`.BaseRouteId
                              AND `UI::TZ`.Id = """ + str(self.projectData['tzId']))
            self.dbOptionRS = self.dbSelect.fetchall()

            optionArr = []
            if self.dbOptionRS:
                for line in self.dbOptionRS:
                    if not line[1] in optionArr:
                        optionArr.append(str(line[1]))
            else:
                optionArr.append('Без опций')
            self.projectData['options'] = optionArr

            tegListArr = []
            if self.radioButton.isChecked():
                for i in self.listWidget.selectedItems():
                    tegListArr.append(str(i.text()))
            elif self.radioButton_2.isChecked():
                tegListArr.append('Без тестовых сборок')
            self.projectData['tegList'] = tegListArr

            if self.projectData['baseRoutes'][0] != 'Не определены':
                if self.projectData['registerChar'][0] != 'Не обнаружен':
                    self.genLayerDict()
                    if self.projectData['maskType'] == 'MLR':
                        self.genPairsDict(self.finalLayerDict)
            self.checkDataForm = CheckDataForm()
            self.checkDataForm.show()

            if self.checkBox.isChecked():
                MyWindow.addMLRTableChk = True
            else:
                MyWindow.addMLRTableChk = False

    def genLayerDict(self):
        dbRequest = """SELECT DISTINCT `MaskLayer`.Id, `MaskLayer`.Position, `MaskLayer`.Layer, `MaskLayer`.Name
                FROM `MaskLayer` LEFT JOIN `MaskLayer_BaseRoute_Relation` 
                ON `MaskLayer`.Id = `MaskLayer_BaseRoute_Relation`.MaskLayerId 
                LEFT JOIN `BaseRoute` ON `MaskLayer_BaseRoute_Relation`.BaseRouteId = `BaseRoute`.Id
                WHERE `BaseRoute`.TechnologyId = """ + str(self.projectData['techId']) + """
                AND `BaseRoute`.Id IN ("""
        for item in self.dbBaseRouteRS[:-1]:
            dbRequest += str(item[0]) + ','
        dbRequest += str(self.dbBaseRouteRS[-1][0]) + ')'

        self.dbSelect.execute(dbRequest)
        self.dbBaseLayerRS = self.dbSelect.fetchall()

        tmpDict = {}
        if self.projectData['options'][0] != 'Без опций':
            for item in self.dbOptionRS:
                if str(item[1]) in tmpDict:
                    tmpDict[str(item[1])] += ', ' + str(item[0])
                else:
                    tmpDict[str(item[1])] = str(item[0])

            optionIdArr = []
            for value in tmpDict.values():
                for item in value.split(', '):
                    optionIdArr.append(str(item))

            dbRequest = """SELECT DISTINCT `MaskLayer`.Id, `MaskLayer`.Position, `MaskLayer`.Layer, `MaskLayer`.Name 
                        FROM `MaskLayer`
                        WHERE `MaskLayer`.Id IN ("""
            for item in optionIdArr[:-1]:
                dbRequest += item + ','
            dbRequest += optionIdArr[-1] + ')'

            self.dbSelect.execute(dbRequest)
            self.dbOptionLayerRS = self.dbSelect.fetchall()

        layerArr = []
        if self.dbBaseLayerRS:
            for item in self.dbBaseLayerRS:
                layerArr.append(str(item[1]) + ';' + str(item[0]))
        if self.dbOptionLayerRS:
            for item in self.dbOptionLayerRS:
                layerArr.append(str(item[1]) + ';' + str(item[0]))
        bubbleSort(layerArr, ';')

        for layer in layerArr:
            dbRequest = """SELECT `MaskLayer`.Id, `MaskLayer`.Layer, `MaskLayer`.Name
                           FROM `MaskLayer`
                           WHERE `MaskLayer`.Id = """ + str(layer.split(';')[1])
            self.dbSelect.execute(dbRequest)
            dbTmpLayerRS = self.dbSelect.fetchall()
            for item in dbTmpLayerRS:
                self.finalLayerDict[str(item[1])] = {'Id': str(item[0]), 'Name': str(item[2])}

    def genPairsDict(self, finalLayersDict):
        idString = ''
        if self.dbBaseLayerRS:
            for item in self.dbBaseLayerRS:
                if item is self.dbBaseLayerRS[0]:
                    idString = str(item[0])
                else:
                    idString += ',' + str(item[0])
        if self.dbOptionLayerRS:
            for item in self.dbOptionLayerRS:
                idString += ',' + str(item[0])

        self.dbSelect.execute("""
                              SELECT * FROM `UI::LayerPairs`
                              WHERE (`UI::LayerPairs`.LayerId1 IN (""" + idString + """)
                              OR `UI::LayerPairs`.LayerId2 IN (""" + idString + "))")
        dbLayerPairsRS = self.dbSelect.fetchall()

        layerPairsMidDict = {}

        topIterNum = 1
        if dbLayerPairsRS:
            for item in dbLayerPairsRS:
                if int(item[2]) > topIterNum:
                    topIterNum = int(item[2])

        for i in range(1, topIterNum + 1):
            for item in dbLayerPairsRS:
                pairFlag = False
                if int(item[2]) == i:
                    layerId1 = int(item[3] or 0)
                    layerId2 = int(item[4] or 0)

                    if layerId1 + layerId2 == layerId1:
                        layer1 = ''
                        for line in self.dbBaseLayerRS:
                            if line[0] == layerId1:
                                layer1 = str(line[2])
                        if layer1 == '':
                            if self.dbOptionLayerRS:
                                for line in self.dbOptionLayerRS:
                                    if line[0] == layerId1:
                                        layer1 = str(line[2])
                        if layer1 != '':
                            if layer1 in finalLayersDict:
                                if not isValueInDict(layerPairsMidDict, layer1):
                                    pairFlag = True
                                    layer2 = '0'
                    elif layerId1 + layerId2 == layerId2:
                        layer2 = ''
                        for line in self.dbBaseLayerRS:
                            if line[0] == layerId1:
                                layer2 = str(line[2])
                        if layer2 == '':
                            if self.dbOptionLayerRS:
                                for line in self.dbOptionLayerRS:
                                    if line[0] == layerId2:
                                        layer2 = str(line[2])
                        if layer2 != '':
                            if layer2 in finalLayersDict:
                                if not isValueInDict(layerPairsMidDict, layer2):
                                    pairFlag = True
                                    layer1 = '0'
                    else:
                        layer1 = ''
                        layer2 = ''
                        for line in self.dbBaseLayerRS:
                            if line[0] == layerId1:
                                layer1 = str(line[2])
                        for line in self.dbBaseLayerRS:
                            if line[0] == layerId2:
                                layer2 = str(line[2])
                        if layer1 == '':
                            if self.dbOptionLayerRS:
                                for line in self.dbOptionLayerRS:
                                    if line[0] == layerId1:
                                        layer1 = str(line[2])
                        if layer2 == '':
                            if self.dbOptionLayerRS:
                                for line in self.dbOptionLayerRS:
                                    if line[0] == layerId2:
                                        layer2 = str(line[2])
                        if not (layer1 == '' or layer2 == ''):
                            if layer1 in finalLayersDict and layer2 in finalLayersDict:
                                if not (isValueInDict(layerPairsMidDict, layer1) or
                                        isValueInDict(layerPairsMidDict, layer2)):
                                    pairFlag = True
                if pairFlag:
                    tmpDict = {'layer1': str(layer1), 'layer2': str(layer2)}
                    if str(layerId1) == '0':
                        tmpDict['layerId1'] = '-'
                    else:
                        tmpDict['layerId1'] = str(layerId1)
                    if str(layerId2) == '0':
                        tmpDict['layerId2'] = '-'
                    else:
                        tmpDict['layerId2'] = str(layerId2)

                    if not str(item[1]) in layerPairsMidDict:
                        layerPairsMidDict[str(item[1])] = tmpDict
                    else:
                        tmpPosition = int(item[1])
                        while str(tmpPosition) in layerPairsMidDict:
                            tmpPosition = int(item[1]) + 1
                        layerPairsMidDict[str(tmpPosition)] = tmpDict

        arrPos = []
        for key in layerPairsMidDict:
            tmpDict = layerPairsMidDict[key]
            layer1 = tmpDict['layer1']
            layer2 = tmpDict['layer2']
            layerId1 = tmpDict['layerId1']
            layerId2 = tmpDict['layerId2']
            arrPos.append(str(key) + ';' + layer1 + ';' + layer2 + ';' + layerId1 + ';' + layerId2)

        bubbleSort(arrPos, ';')
        for i, item in enumerate(arrPos, 1):
            tmpDict = {'layer1': str(item.split(';')[1]), 'layer2': str(item.split(';')[2]),
                       'layerId1': str(item.split(';')[3]), 'layerId2': str(item.split(';')[4])}
            if i <= 9:
                self.finalPairDict['M0' + str(i)] = tmpDict
            else:
                self.finalPairDict['M' + str(i)] = tmpDict


class CheckDataForm(QtWidgets.QWidget, checkdataform.Ui_checkDataForm):

    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.projectData = MyWindow.projectData
        self.finalLayerDict = MyWindow.finalLayerDict
        self.finalPairDict = MyWindow.finalPairDict

        self.ui_projNameLbl.setText(self.projectData['projectName'])
        self.ui_projTypeLbl.setText(self.projectData['projectType'])
        self.ui_tzVerLbl.setText(self.projectData['tzVer'])
        self.ui_maskSetNumLbl.setText(self.projectData['maskSetNum'])
        self.ui_maskTypeLbl.setText(self.projectData['maskType'])
        self.ui_techFamilyLbl.setText(self.projectData['techFamily'] + ' (' + self.projectData['techShortName'] + ')')

        if self.projectData['registerChar'] == 'Не обнаружен':
            self.ui_regTypeLbl.setText(self.projectData['registerChar'])
        else:
            self.ui_regTypeLbl.setText(self.projectData['registerChar'] + ' (' + self.projectData['registerName'] + ')')

        self.ui_baseRouteLbl.setText(cutIntoaPile(self.projectData['baseRoutes'], '\n'))
        self.ui_optionLbl.setText(cutIntoaPile(self.projectData['options'], ', '))
        self.ui_tegListLbl.setText(cutIntoaPile(self.projectData['tegList'], '\n'))
        '''self.ui_baseRouteLbl.clear()
                for item in self.projectData['baseRoutes']:
                    self.ui_baseRouteLbl.setText(self.ui_baseRouteLbl.text() + item + '\n')
        self.ui_optionLbl.setText(self.projectData['options'][0])
        if len(self.projectData['options']) > 1:
            for item in self.projectData['options'][1:]:
                self.ui_optionLbl.setText(self.ui_optionLbl.text() + ', ' + item)
        self.ui_tegListLbl.clear()
        for item in self.projectData['tegList']:
            self.ui_tegListLbl.setText(self.ui_tegListLbl.text() + item + '\n')'''

        self.ui_cancelBtn.clicked.connect(self.close)
        self.ui_okBtn.clicked.connect(self.okBtnClick)

    def okBtnClick(self):
        if self.isCheckFormValid():
            self.createXlList()

    def isCheckFormValid(self):
        errCount = 0

        if self.projectData['baseRoutes'][0] == 'Не определены':
            self.ui_baseRouteLbl.setStyleSheet('background-color: red')
            errCount += 1

        if self.projectData['registerChar'][0] == 'Не обнаружен':
            self.ui_regTypeLbl.setStyleSheet('background-color: red')
            errCount += 1

        # ПРОВЕРКА РЕГУЛЯРНЫМИ ВЫРАЖЕНИЯМИ!

        if not errCount:
            return True
        else:
            return False

    def createXlList(self):
        templatePath = r'C:\Users\Nikita\Desktop\layerList\layerList_template.xlsx'
        wb = openpyxl.load_workbook(templatePath)
        ws = wb['layerList']

        list_projTypeCell = wb.defined_names['list_projTypeCell']
        list_projNameCell = wb.defined_names['list_projNameCell']
        list_maskTypeCell = wb.defined_names['list_maskTypeCell']
        list_tzNumCell = wb.defined_names['list_tzNumCell']
        list_techCell = wb.defined_names['list_techCell']
        list_routesCell = wb.defined_names['list_routesCell']
        list_optsCell = wb.defined_names['list_optsCell']
        list_usedTegsCell = wb.defined_names['list_usedTegsCell']

        ws[next(list_projTypeCell.destinations)[1]].value = MyWindow.projectData['projectType']
        ws[next(list_projNameCell.destinations)[1]].value = MyWindow.projectData['projectName']
        ws[next(list_maskTypeCell.destinations)[1]].value = MyWindow.projectData['maskType']
        ws[next(list_tzNumCell.destinations)[1]].value = MyWindow.projectData['tzVer']
        ws[next(list_techCell.destinations)[1]].value = MyWindow.projectData['techFamily']
        ws[next(list_routesCell.destinations)[1]].value = cutIntoaPile(MyWindow.projectData['baseRoutes'], ', ')
        if MyWindow.projectData['options'][0] != 'Без опций':
            ws[next(list_optsCell.destinations)[1]].value = cutIntoaPile(MyWindow.projectData['options'], ', ')
        ws[next(list_usedTegsCell.destinations)[1]].value = cutIntoaPile(MyWindow.projectData['tegList'], ', ')

        exportDictToFile(self.finalLayerDict, 'finalLayerDict.txt')

        defsexcel.addTableRows(wb, 'list_layerTable', len(self.finalLayerDict) - 1)
        tableRange = openpyxl.worksheet.cell_range.CellRange(defsexcel.getTableObject(ws, 'list_layerTable').ref)

        r = tableRange.min_row + 1
        i = 1
        for layer in self.finalLayerDict:
            tmpDict = self.finalLayerDict[layer]
            barcode = MyWindow.techCode + str(self.projectData['registerChar']) + str(
                self.projectData['maskSetNum'])
            if self.projectData['maskType'] == 'SLR':
                barcode += layer
            elif self.projectData['maskType'] == 'MLR':
                for key in self.finalPairDict:
                    if self.finalPairDict[key]['layer1'] == layer or self.finalPairDict[key]['layer2'] == layer:
                        barcode += key
            barcode += '001'
            ws.cell(row=r, column=1).value = str(i)
            ws.cell(row=r, column=2).value = tmpDict['Name']
            ws.cell(row=r, column=3).value = str(layer)
            ws.cell(row=r, column=4).value = '1'
            ws.cell(row=r, column=5).value = barcode
            ws.cell(row=r, column=6).value = 'V'
            i += 1
            r += 1

        tableRange.min_row = tableRange.min_row + 1
        setTableStyle(ws, tableRange)

        if self.projectData['maskType'] == 'SLR':
            ws._tables.remove(defsexcel.getTableObject(ws, 'list_pairTable'))
            ws._tables.remove(defsexcel.getTableObject(ws, 'list_addPairTable'))  # ОТ СЕБЯ
            ws.delete_cols(8, 5)
        elif self.projectData['maskType'] == 'MLR':
            defsexcel.addTableRows(wb, 'list_pairTable', len(self.finalPairDict) - 1)
            tableRange = openpyxl.worksheet.cell_range.CellRange(defsexcel.getTableObject(ws, 'list_pairTable').ref)
            r = tableRange.min_row
            for pairName in self.finalPairDict:
                tmpDict = self.finalPairDict[pairName]
                ws.cell(row=r, column=8).value = tmpDict['layer1']
                ws.cell(row=r, column=9).value = '+'
                ws.cell(row=r, column=10).value = tmpDict['layer2']
                ws.cell(row=r, column=11).value = pairName
                ws.cell(row=r, column=12).value = '001'
                r += 1
            setTableStyle(ws, tableRange)

            if not MyWindow.addMLRTableChk:
                tableRange = openpyxl.worksheet.cell_range.CellRange(
                    defsexcel.getTableObject(ws, 'list_addPairTable').ref)
                for r in range(tableRange.min_row, tableRange.max_row + 1):
                    for c in range(tableRange.min_col, tableRange.max_col + 1):
                        ws.cell(row=r, column=c).value = ''
                ws._tables.remove(defsexcel.getTableObject(ws, 'list_addPairTable'))

            elif MyWindow.addMLRTableChk:
                tableRange = openpyxl.worksheet.cell_range.CellRange(
                    defsexcel.getTableObject(ws, 'list_addPairTable').ref)
                for r in range(tableRange.min_row, tableRange.max_row + 1):
                    ws.cell(row=r, column=8).value = ''
                    ws.cell(row=r, column=9).value = '+'
                    ws.cell(row=r, column=10).value = ''
                    ws.cell(row=r, column=11).value = ''
                    ws.cell(row=r, column=12).value = ''
                setTableStyle(ws, tableRange)

        templatePath = r'C:\Users\Nikita\Desktop\layerList' + '\\' + self.projectData['projectName'] + '_List_' + \
                       self.projectData['tzVer'] + '.xlsx'
        wb.save(templatePath)


def cutIntoaPile(arr, delimiter):
    tmpString = ''
    for item in arr:
        if not tmpString:
            tmpString = str(item)
        else:
            tmpString = tmpString + delimiter + item
    return tmpString


def bubbleSort(inputArr, delimiter):
    for i in range(len(inputArr) - 1):
        for j in range(len(inputArr) - i - 1):
            if int(inputArr[j].split(delimiter)[0]) > int(inputArr[j + 1].split(delimiter)[0]):
                inputArr[j], inputArr[j + 1] = inputArr[j + 1], inputArr[j]
            elif int(inputArr[j].split(delimiter)[0]) == int(inputArr[j + 1].split(delimiter)[0]):
                if int(inputArr[j].split(delimiter)[1]) > int(inputArr[j + 1].split(delimiter)[1]):
                    inputArr[j], inputArr[j + 1] = inputArr[j + 1], inputArr[j]


def getDictValues(inputDict, depth=0):
    outStr = ''
    for key in inputDict:
        if not isinstance(inputDict[key], dict):
            pass
            '''for i in range(depth):
                outStr += '    '
            for i in range(depth):'''
        else:
            outStr = outStr + '\n'
            for i in range(depth):
                outStr = outStr + '    '
            for i in range(depth):
                outStr = outStr + '-'
            outStr = outStr + ' ' + str(key) + '[dict]'
            outStr = outStr + '\n' + getDictValues(inputDict[key], depth + 1)
    return outStr


def exportDictToFile(inputDict, fileName):
    templatePath = r'C:\Users\Nikita\Desktop\layerList\Folder_test'
    if not os.path.isdir(templatePath):
        os.mkdir(templatePath)
    templatePath += r'\layerList_debug'
    if not os.path.isdir(templatePath):
        os.mkdir(templatePath)

    debugLog = open(templatePath + '\\' + fileName, 'w')
    debugLog.write(getDictValues(inputDict))
    debugLog.close()


def isValueInDict(dict, searchVal):
    if len(dict) != 0:
        for key in dict:
            tmpDict = dict[key]
            for key2 in tmpDict:
                if tmpDict[key2] == searchVal:
                    return True
    return False


def setTableStyle(ws, tableRange):
    ft = openpyxl.styles.Font(name='Calibri', size=10, bold=False)
    al = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    brd = openpyxl.styles.Border(left=None, right=None, top=None, bottom=None)
    for r in range(tableRange.min_row, tableRange.max_row + 1):
        ws.row_dimensions[r].height = 12
        for c in range(tableRange.min_col, tableRange.max_col + 1):
            ws.cell(row=r, column=c).font = ft
            ws.cell(row=r, column=c).alignment = al
            ws.cell(row=r, column=c).border = brd


app = QtWidgets.QApplication([])


application = MyWindow()
application.show()

sys.exit(app.exec())
